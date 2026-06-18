# -*- coding: utf-8 -*-
"""
Monotributo — recategorización.

Lee una planilla de facturación (la del estudio: hoja 'ACUMULADOS' con los totales
por mes, o las hojas mensuales) y, con la escala vigente, calcula para un período:
  - facturación acumulada de los últimos 12 meses,
  - categoría que le corresponde,
  - cuánto más puede facturar para MANTENERSE en la categoría,
  - cuánto le falta para PASARSE a la siguiente (o quedar excluido).
"""

import io
import json
import os
from datetime import date

_MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
_MESNOM = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
           "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def cargar_escala(ruta=None):
    if ruta is None:
        ruta = os.path.join(os.path.dirname(__file__), "escala_monotributo.json")
    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


# --------------------------------------------------------------------------
# Lectura de la facturación mensual de la planilla
# --------------------------------------------------------------------------
def leer_facturacion(data):
    """Devuelve un dict {(año, mes): monto} con la facturación por mes.

    Primero intenta la hoja 'ACUMULADOS' (meses en filas, años en columnas);
    si no, suma la columna MONTO de las hojas mensuales (nombre tipo 'M-AAAA')."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    serie = _desde_acumulados(wb)
    if serie:
        return serie
    return _desde_hojas_mensuales(wb)


def _desde_acumulados(wb):
    hoja = next((h for h in wb.sheetnames if "acumulad" in h.lower()), None)
    if hoja is None:
        return {}
    ws = wb[hoja]
    # Buscar la fila de años (celdas que son 20xx) y las columnas correspondientes.
    años_col = {}
    fila_años = None
    for r in range(1, min(ws.max_row, 6) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            try:
                n = int(v)
            except (TypeError, ValueError):
                continue
            if 2018 <= n <= 2100:
                años_col[c] = n
                fila_años = r
        if años_col:
            break
    if not años_col:
        return {}
    serie = {}
    for r in range(fila_años + 1, ws.max_row + 1):
        etiqueta = str(ws.cell(r, 1).value or "").strip().lower()
        mes = _MESES.get(etiqueta)
        if not mes:
            continue
        for c, año in años_col.items():
            monto = _num(ws.cell(r, c).value)
            if monto:
                serie[(año, mes)] = serie.get((año, mes), 0.0) + monto
    return serie


def _desde_hojas_mensuales(wb):
    import re
    serie = {}
    for nm in wb.sheetnames:
        m = re.match(r"^\s*(\d{1,2})\s*[-/]\s*(\d{2,4})\s*$", nm)
        if not m:
            continue
        mes = int(m.group(1))
        año = int(m.group(2))
        if año < 100:
            año += 2000
        if not (1 <= mes <= 12):
            continue
        ws = wb[nm]
        # buscar columna MONTO
        c_monto = None
        for r in range(1, min(ws.max_row, 5) + 1):
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(r, c).value or "").strip().lower() == "monto":
                    c_monto = c
                    fila_ini = r + 1
                    break
            if c_monto:
                break
        if not c_monto:
            continue
        total = sum(_num(ws.cell(r, c_monto).value) for r in range(fila_ini, ws.max_row + 1))
        serie[(año, mes)] = serie.get((año, mes), 0.0) + total
    return serie


def meses_disponibles(serie):
    """Lista de (año, mes) con datos, ordenada."""
    return sorted(k for k, v in serie.items() if v)


def acumulado_rango(serie, desde_año, desde_mes, hasta_año, hasta_mes):
    """Suma los meses entre (desde) y (hasta) inclusive. Devuelve (total, cant_meses, detalle)."""
    total = 0.0
    detalle = []
    y, m = desde_año, desde_mes
    while (y, m) <= (hasta_año, hasta_mes):
        monto = serie.get((y, m), 0.0)
        total += monto
        detalle.append((y, m, monto))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return round(total, 2), len(detalle), detalle


def anualizar(total, meses):
    """Proyecta a 12 meses (para inscriptos hace menos de 1 año)."""
    if not meses:
        return 0.0
    return round(total / meses * 12, 2)


def acumulado_12m(serie, hasta_año, hasta_mes):
    """Suma los 12 meses que terminan en (hasta_año, hasta_mes)."""
    total = 0.0
    detalle = []
    y, m = hasta_año, hasta_mes
    for _ in range(12):
        monto = serie.get((y, m), 0.0)
        total += monto
        detalle.append((y, m, monto))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return round(total, 2), list(reversed(detalle))


# --------------------------------------------------------------------------
# Análisis de categoría
# --------------------------------------------------------------------------
def analizar(acumulado, actividad, escala):
    """Devuelve el análisis de categoría para un acumulado de 12 meses."""
    tabla = escala.get(actividad) or escala.get("servicios")
    cat = None
    for i, fila in enumerate(tabla):
        if acumulado <= fila["tope"]:
            cat = i
            break

    if cat is None:  # supera el tope máximo (K) -> excluido
        ult = tabla[-1]
        return {
            "categoria": "EXCLUIDO",
            "cuota": None,
            "tope_categoria": ult["tope"],
            "acumulado": acumulado,
            "margen_mantenerse": 0.0,
            "excede_por": round(acumulado - ult["tope"], 2),
            "siguiente": None,
            "tope_siguiente": None,
            "vigencia": escala.get("vigencia", ""),
        }

    fila = tabla[cat]
    sig = tabla[cat + 1] if cat + 1 < len(tabla) else None
    margen = round(fila["tope"] - acumulado, 2)
    return {
        "categoria": fila["cat"],
        "cuota": fila.get("cuota"),
        "tope_categoria": fila["tope"],
        "acumulado": acumulado,
        "margen_mantenerse": margen,            # cuánto más puede facturar sin pasar
        "para_pasarse": margen,                 # si factura más que esto, sube
        "siguiente": sig["cat"] if sig else "EXCLUIDO",
        "tope_siguiente": sig["tope"] if sig else None,
        "cuota_siguiente": sig.get("cuota") if sig else None,
        "vigencia": escala.get("vigencia", ""),
    }
