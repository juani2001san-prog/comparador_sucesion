# -*- coding: utf-8 -*-
"""
Logica pura para transformar un JWIN (export de asientos) en el archivo .ps3
de ancho fijo que se sube al sistema contable.

No depende de Streamlit: se puede testear e importar desde cualquier lado.

Regla del renglon .ps3 (68 caracteres, validada contra archivos reales):
    tipo_registro(1) + cuenta_CO(8) + fecha_YYYYMMDD(8) + tipo_asiento(1)
    + nro_asiento_interno(6) + nro_renglon(3) + detalle(25)
    + signo(1) + importe_centavos(15)

El codigo de cuenta del sistema (CO, 8 digitos) son los primeros 8 digitos
del codigo de cuenta del JWIN (9 digitos). La validacion de "cuenta inexistente"
se hace a nivel CO (8 digitos) contra el plan de cuentas.
"""

import io
import json
import os
from dataclasses import dataclass, field
from datetime import datetime, date

import openpyxl

# Epoca de Excel para convertir fechas a numero de serie (igual que el *-ps3)
_EXCEL_EPOCH = date(1899, 12, 30)

# Encabezado de la hoja DIARIO, identico a los archivos *-ps3 originales
DIARIO_HEADER = [
    " ", "Cód. cuenta", None, None, None, None, "DETALLE", None, None, None,
    None, "IMPORTE", "IMPORTE", None, "NRO. ASIENTO INTERNO", "FECHA ASIENTO",
    None, "CÓD. CUENTA", None,
]

# --- Constantes del formato (confirmadas con los datos reales) ---
TIPO_REGISTRO = "1"
TIPO_ASIENTO = "2"
ANCHO_DETALLE = 25
ANCHO_IMPORTE = 15
LARGO_LINEA = 68
FIN_LINEA = "\r\n"  # el .ps3 usa CRLF y deja salto al final

# Columnas del JWIN (1-indexadas, segun el export estandar)
COL_FECHA = 1            # FECHA ASIENTO
COL_ASIENTO_INTERNO = 4  # NRO. ASIENTO INTERNO
COL_RENGLON = 6          # NRO. RENGLON
COL_CUENTA = 7           # COD. CUENTA (9 digitos)
COL_DEBE = 9             # DEBE
COL_HABER = 10           # HABER
COL_LEYENDA = 12         # se usa para armar el detalle (LEFT 3 + MID 13,7)
COL_PROVEEDOR = 13       # se usa para armar el detalle (LEFT 14)


# --------------------------------------------------------------------------
# Carga del plan de cuentas fijo
# --------------------------------------------------------------------------
def cargar_plan(ruta_json=None):
    """Devuelve (co_validos: dict CO8->nombre, correcciones: dict cod9->cod9)."""
    if ruta_json is None:
        ruta_json = os.path.join(os.path.dirname(__file__), "plan_cuentas_microenv.json")
    with open(ruta_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    co_validos = {}
    for c in data["cuentas"]:
        cod = str(c["codigo"]).strip()
        co8 = cod[:8]
        # el primero gana: en el plan los codigos van de lo general a lo particular
        co_validos.setdefault(co8, c.get("detalle", ""))
    correcciones = {str(k): str(v) for k, v in data.get("correcciones", {}).items()}
    return co_validos, correcciones


# --------------------------------------------------------------------------
# Helpers de transformacion
# --------------------------------------------------------------------------
def _texto(valor):
    return "" if valor is None else str(valor)


def codigo_str(valor):
    """Normaliza el codigo de cuenta del JWIN a string de digitos."""
    if valor is None:
        return ""
    if isinstance(valor, float):
        return str(int(valor))
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip()


def armar_detalle(leyenda, proveedor):
    """Replica la formula del JWIN: LEFT(L,3)+' '+MID(L,13,7)+LEFT(M,14).

    Devuelve el detalle SIN rellenar (como va en la columna DETALLE del Excel).
    Para el renglon .ps3 se rellena luego a 25 con espacios.
    """
    leyenda = _texto(leyenda)
    proveedor = _texto(proveedor)
    s = leyenda[:3] + " " + leyenda[12:19] + proveedor[:14]
    return s[:ANCHO_DETALLE]


def fecha_serial(valor):
    """Convierte la fecha a numero de serie de Excel (como guarda el *-ps3)."""
    if isinstance(valor, datetime):
        valor = valor.date()
    if isinstance(valor, date):
        return (valor - _EXCEL_EPOCH).days
    s = fecha_str(valor)
    if len(s) == 8:
        d = datetime.strptime(s, "%Y%m%d").date()
        return (d - _EXCEL_EPOCH).days
    return None


def fecha_str(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y%m%d")
    # por las dudas: texto tipo 2026-02-02 o serial
    s = _texto(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return ""


def _num(valor):
    if valor is None or valor == "":
        return 0.0
    return float(valor)


def _leer_filas(origen):
    """Lee el JWIN y devuelve una lista de filas (cada fila = lista de valores).

    Soporta los dos formatos:
      - .xlsx nuevo (es un ZIP, empieza con 'PK')  -> openpyxl
      - .xls viejo (formato OLE2 de Excel 97-2003) -> xlrd
    Algunos archivos tienen extension .xls pero por dentro son .xlsx: por eso
    se detecta por el contenido (los primeros bytes), no por la extension.
    Las fechas se devuelven como datetime en ambos casos.
    """
    data = origen.read() if hasattr(origen, "read") else open(origen, "rb").read()

    # 'PK' = archivo ZIP = xlsx (aunque tenga extension .xls)
    if data[:2] == b"PK":
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        return [list(fila) for fila in ws.iter_rows(values_only=True)]

    # Si no, es un .xls viejo -> xlrd
    import xlrd  # import diferido: solo se necesita para .xls viejos
    wb = xlrd.open_workbook(file_contents=data)
    sh = wb.sheet_by_index(0)
    filas = []
    for r in range(sh.nrows):
        fila = []
        for c in range(sh.ncols):
            celda = sh.cell(r, c)
            # xlrd entrega las fechas como numero de serie; las paso a datetime
            if celda.ctype == xlrd.XL_CELL_DATE:
                fila.append(xlrd.xldate.xldate_as_datetime(celda.value, wb.datemode))
            else:
                fila.append(celda.value)
        filas.append(fila)
    return filas


# --------------------------------------------------------------------------
# Resultado del procesamiento
# --------------------------------------------------------------------------
@dataclass
class Resultado:
    empresa: str
    lineas: list = field(default_factory=list)          # renglones .ps3 (str de 68)
    filas: list = field(default_factory=list)           # filas del DIARIO (listas de 19 valores)
    total_debe: float = 0.0
    total_haber: float = 0.0
    cant_renglones: int = 0
    cant_asientos: int = 0
    correcciones_aplicadas: dict = field(default_factory=dict)  # cod9 -> (cod9_nuevo, veces)
    inexistentes: dict = field(default_factory=dict)    # cod9 -> {'co':co8,'veces':n,'filas':[...]}
    periodo: str = ""                                   # MM-YYYY mas frecuente
    errores_largo: list = field(default_factory=list)   # filas cuyo renglon no dio 68 chars

    @property
    def balanceado(self):
        return round(self.total_debe - self.total_haber, 2) == 0

    @property
    def diferencia(self):
        return round(self.total_debe - self.total_haber, 2)

    def texto_ps3(self):
        return "".join(linea + FIN_LINEA for linea in self.lineas)


# --------------------------------------------------------------------------
# Procesamiento principal
# --------------------------------------------------------------------------
def procesar_jwin(origen, empresa, co_validos, correcciones):
    """
    origen: ruta o file-like (bytes) del .xlsx JWIN.
    empresa: etiqueta ('Micro' / 'Pruebas') para nombrar la salida.
    co_validos: dict CO8 -> nombre (cuentas validas).
    correcciones: dict cod9 -> cod9 (reemplazos a aplicar antes de mapear).

    Devuelve un Resultado. No frena ante inexistentes: las registra para que
    la capa de UI decida (el usuario tipea la correccion a mano).
    """
    filas = _leer_filas(origen)

    res = Resultado(empresa=empresa)
    asientos = set()
    periodos = {}

    def celda(fila, col):
        """Valor de una columna 1-indexada; None si la fila no llega hasta ahi."""
        idx = col - 1
        return fila[idx] if 0 <= idx < len(fila) else None

    # saltar la fila de encabezado (fila 1); r conserva el numero de fila real
    for r, fila in enumerate(filas[1:], start=2):
        cod9 = codigo_str(celda(fila, COL_CUENTA))
        if not cod9:
            continue  # fila vacia / sin cuenta

        # correccion conocida a nivel codigo del JWIN
        if cod9 in correcciones:
            nuevo = correcciones[cod9]
            prev = res.correcciones_aplicadas.get(cod9)
            res.correcciones_aplicadas[cod9] = (nuevo, (prev[1] + 1) if prev else 1)
            cod9 = nuevo

        co8 = cod9[:8]

        fecha = celda(fila, COL_FECHA)
        debe = _num(celda(fila, COL_DEBE))
        haber = _num(celda(fila, COL_HABER))
        asiento_interno = int(_num(celda(fila, COL_ASIENTO_INTERNO)))
        renglon = int(_num(celda(fila, COL_RENGLON)))
        detalle = armar_detalle(celda(fila, COL_LEYENDA), celda(fila, COL_PROVEEDOR))

        # registrar cuenta inexistente (no frena el proceso)
        if co8 not in co_validos:
            info = res.inexistentes.setdefault(cod9, {"co": co8, "veces": 0, "filas": []})
            info["veces"] += 1
            if len(info["filas"]) < 20:
                info["filas"].append(r)

        # signo e importe
        neto = round(debe - haber, 2)
        signo = "0" if neto >= 0 else "1"
        importe = int(round(abs(neto) * 100))
        detalle_25 = detalle.ljust(ANCHO_DETALLE)[:ANCHO_DETALLE]
        serial = fecha_serial(fecha)

        linea = (
            TIPO_REGISTRO
            + co8.zfill(8)
            + fecha_str(fecha)
            + TIPO_ASIENTO
            + f"{asiento_interno:06d}"
            + f"{renglon:03d}"
            + detalle_25
            + signo
            + f"{importe:0{ANCHO_IMPORTE}d}"
        )
        if len(linea) != LARGO_LINEA:
            res.errores_largo.append((r, len(linea), linea))

        # fila del DIARIO (19 columnas, identicas al *-ps3)
        res.filas.append([
            int(TIPO_REGISTRO),          # A  tipo registro
            int(co8),                    # B  Cod. cuenta (CO 8 dig)
            serial,                      # C  fecha (serie Excel)
            int(TIPO_ASIENTO),           # D  tipo asiento
            asiento_interno,             # E  nro asiento interno
            renglon,                     # F  nro renglon
            detalle,                     # G  DETALLE (sin rellenar)
            int(signo),                  # H  signo (0 debe / 1 haber)
            importe,                     # I  importe en centavos
            None,                        # J  (vacio)
            None,                        # K  (vacio)
            neto,                        # L  IMPORTE con signo
            abs(neto),                   # M  IMPORTE absoluto
            importe,                     # N  importe en centavos
            asiento_interno,             # O  nro asiento interno
            serial,                      # P  fecha (serie Excel)
            None,                        # Q  (vacio)
            cod9,                        # R  COD. CUENTA (9 dig, texto)
            linea,                       # S  renglon .ps3
        ])

        res.lineas.append(linea)
        res.total_debe += debe
        res.total_haber += haber
        asientos.add(asiento_interno)

        ym = fecha_str(fecha)
        if len(ym) == 8:
            clave = ym[4:6] + "-" + ym[0:4]
            periodos[clave] = periodos.get(clave, 0) + 1

    res.cant_renglones = len(res.lineas)
    res.cant_asientos = len(asientos)
    res.total_debe = round(res.total_debe, 2)
    res.total_haber = round(res.total_haber, 2)
    if periodos:
        res.periodo = max(periodos, key=periodos.get)
    return res


def nombre_archivo_ps3(empresa, periodo):
    """Micro-02-2026.ps3 / Prueba-02-2026.ps3 (ancho fijo para subir a JWIN)."""
    etiqueta = "Micro" if empresa.lower().startswith("micro") else "Prueba"
    return f"{etiqueta}-{periodo}.ps3" if periodo else f"{etiqueta}.ps3"


def nombre_archivo_xlsx(empresa, periodo):
    """MICRO-ps3-02-2026.xlsx / PRUEBAS-ps3-02-2026.xlsx."""
    etiqueta = "MICRO" if empresa.lower().startswith("micro") else "PRUEBAS"
    return f"{etiqueta}-ps3-{periodo}.xlsx" if periodo else f"{etiqueta}-ps3.xlsx"


def construir_diario_xlsx(resultado):
    """Devuelve los bytes de un .xlsx con la hoja DIARIO (19 columnas),
    igual a los archivos *-ps3 originales. La columna S es el renglon .ps3."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DIARIO"
    ws.append(DIARIO_HEADER)
    # formato de fecha para las columnas C (3) y P (16), igual al *-ps3
    for fila in resultado.filas:
        ws.append(fila)
    # mostrar C y P como fecha dd/mm/aaaa (estan guardadas como serie Excel)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 3).number_format = "dd/mm/yyyy"
        ws.cell(r, 16).number_format = "dd/mm/yyyy"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
