# -*- coding: utf-8 -*-
"""
Importación AFIP → JWIN.

Toma el CSV de AFIP ("Mis Comprobantes - Comprobantes Recibidos") y le agrega
la columna **Rubro** buscando el CUIT del emisor en una lista maestra de
proveedores (CUIT → código de rubro 1..17). Devuelve el archivo listo para
importar a JWIN (Excel y/o CSV) y la lista de proveedores nuevos sin rubro.

El CSV de AFIP viene con separador ';', decimales con coma y, a veces, comillas.
Los valores se preservan tal cual (no se reconvierten) para no alterar el formato
que espera JWIN; solo se agrega el rubro al final.
"""

import csv
import io
from datetime import date, datetime


# --------------------------------------------------------------------------
# Lectura del CSV de AFIP
# --------------------------------------------------------------------------
def _decodificar(data):
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1", "replace")


def _leer_csv_afip(data):
    texto = _decodificar(data)
    # autodetectar separador (AFIP usa ';', pero por las dudas)
    sep = ";" if texto.count(";") >= texto.count(",") else ","
    filas = list(csv.reader(io.StringIO(texto), delimiter=sep))
    return [f for f in filas if any(str(c).strip() for c in f)]


def _celda_texto(v):
    """Convierte una celda de Excel al texto estilo AFIP (fecha ISO, coma decimal)."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return f"{v:.2f}".replace(".", ",")
    return str(v).strip()


def _leer_excel_afip(data):
    """Lee el Excel de AFIP: busca la hoja con el encabezado correcto y devuelve
    filas (encabezado + datos) como texto, ignorando una columna 'Rubro' previa."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    hoja = None
    for nm in wb.sheetnames:
        ws = wb[nm]
        cab = " ".join(str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)).lower()
        if "nro. doc. emisor" in cab or "fecha de emisi" in cab:
            hoja = ws
            break
    if hoja is None:
        hoja = wb.active
    filas = []
    for fila in hoja.iter_rows(values_only=True):
        if any(c not in (None, "") for c in fila):
            filas.append([_celda_texto(c) for c in fila])
    return filas


def _leer_entrada(data):
    """Acepta CSV o Excel (xlsx/xls). Devuelve filas (encabezado + datos) como texto."""
    if data[:2] == b"PK" or data[:4] == b"\xd0\xcf\x11\xe0":
        return _leer_excel_afip(data)
    return _leer_csv_afip(data)


# Layout canónico que espera el importador de JWIN (formato AFIP "viejo", 30 cols).
_CANON = [
    "Fecha de Emisión", "Tipo de Comprobante", "Punto de Venta", "Número Desde",
    "Número Hasta", "Cód. Autorización", "Tipo Doc. Emisor", "Nro. Doc. Emisor",
    "Denominación Emisor", "Tipo Doc. Receptor", "Nro. Doc. Receptor", "Tipo Cambio",
    "Moneda", "Imp. Neto Gravado IVA 0%", "IVA 2,5%", "Imp. Neto Gravado IVA 2,5%",
    "IVA 5%", "Imp. Neto Gravado IVA 5%", "IVA 10,5%", "Imp. Neto Gravado IVA 10,5%",
    "IVA 21%", "Imp. Neto Gravado IVA 21%", "IVA 27%", "Imp. Neto Gravado IVA 27%",
    "Imp. Neto Gravado Total", "Imp. Neto No Gravado", "Imp. Op. Exentas",
    "Otros Tributos", "Total IVA", "Imp. Total",
]
# Nombres equivalentes en el formato AFIP "nuevo" (montos en pesos / Vendedor).
_ALT = {
    "Número Desde": ["Número de Comprobante"], "Número Hasta": ["Número de Comprobante"],
    "Tipo Doc. Emisor": ["Tipo Doc. Vendedor"], "Nro. Doc. Emisor": ["Nro. Doc. Vendedor"],
    "Denominación Emisor": ["Denominación Vendedor"], "Tipo Cambio": ["Tipo de Cambio"],
    "Moneda": ["Moneda Original"], "Imp. Neto Gravado IVA 0%": ["Neto Gravado IVA 0%"],
    "IVA 2,5%": ["Importe IVA 2,5%"], "Imp. Neto Gravado IVA 2,5%": ["Neto Gravado IVA 2,5%"],
    "IVA 5%": ["Importe IVA 5%"], "Imp. Neto Gravado IVA 5%": ["Neto Gravado IVA 5%"],
    "IVA 10,5%": ["Importe IVA 10,5%"], "Imp. Neto Gravado IVA 10,5%": ["Neto Gravado IVA 10,5%"],
    "IVA 21%": ["Importe IVA 21%"], "Imp. Neto Gravado IVA 21%": ["Neto Gravado IVA 21%"],
    "IVA 27%": ["Importe IVA 27%"], "Imp. Neto Gravado IVA 27%": ["Neto Gravado IVA 27%"],
    "Imp. Neto Gravado Total": ["Total Neto Gravado"], "Imp. Neto No Gravado": ["Importe No Gravado"],
    "Imp. Op. Exentas": ["Importe Exento"],
    "Otros Tributos": ["Importe de Percepciones de Ingresos Brutos"],
    "Imp. Total": ["Importe Total"],
}


def _norm_h(s):
    return " ".join(str(s).strip().lower().split())


def _a_layout_jwin(encab, datos):
    """Reordena cualquier formato de AFIP al layout canónico que espera JWIN.
    Para el formato viejo es identidad; el nuevo lo remapea por significado."""
    idx = {}
    for i, h in enumerate(encab):
        idx.setdefault(_norm_h(h), i)

    def buscar(canon):
        for c in [canon] + _ALT.get(canon, []):
            j = idx.get(_norm_h(c))
            if j is not None:
                return j
        return None

    cols = [buscar(c) for c in _CANON]
    nuevos = [[(f[j] if (j is not None and j < len(f)) else "") for j in cols] for f in datos]
    return list(_CANON), nuevos


def _norm_cuit(valor):
    if valor is None:
        return ""
    s = str(valor).strip()
    if s.endswith(".0"):  # por si vino como número
        s = s[:-2]
    return "".join(ch for ch in s if ch.isdigit())


# --------------------------------------------------------------------------
# Lectura del maestro (Proveedores + Rubros)
# --------------------------------------------------------------------------
def _leer_maestro(data):
    """Devuelve (proveedores: CUIT->rubro, rubros: cod->descripcion)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    proveedores = {}
    hoja_prov = next((h for h in wb.sheetnames if "proveedor" in h.lower()), None)
    if hoja_prov:
        ws = wb[hoja_prov]
        hdr = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

        def idx(clave):
            for i, h in enumerate(hdr):
                if clave in h:
                    return i + 1
            return None

        c_cuit = idx("cuit") or 1
        c_rubro = idx("rubro") or 3
        for r in range(2, ws.max_row + 1):
            cuit = _norm_cuit(ws.cell(r, c_cuit).value)
            rub = ws.cell(r, c_rubro).value
            if cuit and rub not in (None, ""):
                try:
                    proveedores[cuit] = int(rub)
                except (TypeError, ValueError):
                    proveedores[cuit] = rub

    rubros = {}
    hoja_rub = next((h for h in wb.sheetnames if "rubro" in h.lower()), None)
    if hoja_rub:
        ws = wb[hoja_rub]
        for r in range(2, ws.max_row + 1):
            cod, desc = ws.cell(r, 1).value, ws.cell(r, 2).value
            if cod not in (None, ""):
                try:
                    rubros[int(cod)] = desc
                except (TypeError, ValueError):
                    pass
    return proveedores, rubros


# --------------------------------------------------------------------------
# Procesamiento
# --------------------------------------------------------------------------
def procesar(csv_bytes, maestro_bytes, extra=None):
    """Devuelve (encabezado, filas_con_rubro, desconocidos, stats, rubros).

    `extra`: dict CUIT->rubro con asignaciones cargadas a mano en la app
    (proveedores nuevos que todavía no están en el maestro).
    """
    filas = _leer_entrada(csv_bytes)
    if not filas:
        raise ValueError("El archivo de AFIP está vacío.")
    encab, datos = list(filas[0]), [list(f) for f in filas[1:]]

    # Si ya traía una columna 'Rubro' al final (ej. un Excel ya armado), la saco.
    if encab and str(encab[-1]).strip().lower() == "rubro":
        encab = encab[:-1]
        datos = [f[:len(encab)] for f in datos]

    # Dejamos el archivo tal cual viene de AFIP y solo agregamos el Rubro al final.
    def col(*nombres):
        for nm in nombres:
            for i, h in enumerate(encab):
                if nm.lower() in str(h).strip().lower():
                    return i
        return None

    # Acepta el formato viejo ('Emisor') y el nuevo 'montos en pesos' ('Vendedor').
    i_cuit = col("Nro. Doc. Emisor", "Nro. Doc. Vendedor", "Nro. Doc")
    i_deno = col("Denominación Emisor", "Denominación Vendedor", "Denominación")
    if i_cuit is None:
        raise ValueError("No encontré la columna del CUIT del proveedor "
                         "('Nro. Doc. Emisor' o 'Nro. Doc. Vendedor') en el archivo de AFIP.")

    proveedores, rubros = _leer_maestro(maestro_bytes)
    # Sumar las asignaciones cargadas a mano (tienen prioridad).
    for cuit, rub in (extra or {}).items():
        c = _norm_cuit(cuit)
        if c and rub not in (None, ""):
            try:
                proveedores[c] = int(rub)
            except (TypeError, ValueError):
                proveedores[c] = rub

    salida = []
    desconocidos = {}
    asignados = 0
    for fila in datos:
        cuit = _norm_cuit(fila[i_cuit]) if i_cuit < len(fila) else ""
        rub = proveedores.get(cuit, "")
        if rub == "":
            deno = fila[i_deno] if (i_deno is not None and i_deno < len(fila)) else ""
            if cuit:
                desconocidos.setdefault(cuit, deno)
        else:
            asignados += 1
        salida.append(list(fila) + [rub])

    encab_out = list(encab) + ["Rubro"]
    stats = {
        "comprobantes": len(datos),
        "asignados": asignados,
        "sin_rubro": len(datos) - asignados,
        "proveedores_nuevos": len(desconocidos),
    }
    return encab_out, salida, desconocidos, stats, rubros


# --------------------------------------------------------------------------
# Salidas
# --------------------------------------------------------------------------
# Catálogo de rubros estándar (JWIN). Sirve para armar una plantilla nueva.
RUBROS_ESTANDAR = [
    (1, "Movilidad y Viaticos"), (2, "Fletes"), (3, "Honorarios y Aranceles"),
    (4, "Repuestos y reparaciones"), (5, "Combustibles"), (6, "Internet"),
    (7, "Obra Social"), (8, "Semillas e insumos"), (9, "Telefonia"),
    (10, "Agroquimicos"), (11, "Gastos Varios"), (12, "Hacienda"),
    (13, "Gastos 10.5"), (14, "Laboreos"), (15, "Luz, gas y agua"),
    (16, "Alquileres"), (17, "Combustibles 10.5"),
]


def construir_plantilla_maestro(rubros=None):
    """Devuelve un Excel maestro vacío: hoja Rubros (catálogo) + hoja Proveedores
    (solo encabezados) + Instrucciones. Para empezar de cero un maestro."""
    import openpyxl
    from openpyxl.styles import Font

    rubros = rubros or RUBROS_ESTANDAR
    bold = Font(bold=True)
    wb = openpyxl.Workbook()

    wi = wb.active
    wi.title = "Instrucciones"
    txt = [
        "EXCEL MAESTRO DE PROVEEDORES (para AFIP -> JWIN)",
        "",
        "Hoja 'Rubros': catálogo de rubros (código 1..17). Ya viene cargado.",
        "Hoja 'Proveedores': cargá acá cada proveedor con su CUIT y el código de rubro.",
        "   - CUIT: los 11 dígitos, sin guiones.",
        "   - Rubro: el número del catálogo (mirá la hoja Rubros).",
        "",
        "Tip: en la herramienta AFIP -> JWIN, cuando aparezca un proveedor nuevo,",
        "lo cargás ahí mismo y la app te devuelve este maestro ya actualizado.",
    ]
    for i, t in enumerate(txt, start=1):
        wi.cell(i, 1, t)

    wr = wb.create_sheet("Rubros")
    wr.cell(1, 1, "Código").font = bold
    wr.cell(1, 2, "Descripción").font = bold
    for i, (cod, desc) in enumerate(rubros, start=2):
        wr.cell(i, 1, cod)
        wr.cell(i, 2, desc)

    wp = wb.create_sheet("Proveedores")
    for j, t in enumerate(["CUIT", "Razón Social", "Rubro", "Descripción Rubro"], start=1):
        wp.cell(1, j, t).font = bold
    wp.column_dimensions["A"].width = 16
    wp.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_maestro_actualizado(maestro_bytes, nuevos):
    """Devuelve el Excel maestro con los proveedores nuevos AGREGADOS al final
    de la hoja Proveedores. `nuevos` = lista de (cuit, denominacion, rubro, desc)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(maestro_bytes))
    hoja = next((h for h in wb.sheetnames if "proveedor" in h.lower()), None)
    if hoja is None:
        hoja = "Proveedores"
        ws = wb.create_sheet(hoja)
        ws.append(["CUIT", "Razón Social", "Rubro", "Descripción Rubro"])
    ws = wb[hoja]
    fila = ws.max_row + 1
    for cuit, deno, rubro, desc in nuevos:
        ws.cell(fila, 1, str(cuit))
        ws.cell(fila, 2, deno)
        ws.cell(fila, 3, rubro)
        ws.cell(fila, 4, desc)
        fila += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_csv(encab, filas):
    """CSV UTF-8 con separador ';' (mismo formato que AFIP) + columna Rubro."""
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    w.writerow(encab)
    for f in filas:
        w.writerow(f)
    return buf.getvalue().encode("utf-8-sig")


def construir_excel(encab, filas, desconocidos, rubros):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AFIP"
    bold = Font(bold=True)
    rojo = PatternFill("solid", fgColor="FFC7CE")

    for j, t in enumerate(encab, start=1):
        ws.cell(1, j, t).font = bold
    col_rubro = len(encab)
    for i, fila in enumerate(filas, start=2):
        for j, v in enumerate(fila, start=1):
            c = ws.cell(i, j, v)
            if j == col_rubro and (v == "" or v is None):
                c.fill = rojo  # sin rubro: resaltado
    ws.freeze_panes = "A2"

    # Hoja de proveedores nuevos sin rubro
    wd = wb.create_sheet("Proveedores nuevos")
    wd.cell(1, 1, "CUIT").font = bold
    wd.cell(1, 2, "Denominación").font = bold
    wd.cell(1, 3, "Rubro (completar)").font = bold
    for i, (cuit, deno) in enumerate(sorted(desconocidos.items()), start=2):
        wd.cell(i, 1, cuit)
        wd.cell(i, 2, deno)
    wd.column_dimensions["A"].width = 16
    wd.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
