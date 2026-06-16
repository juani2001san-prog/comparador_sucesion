# -*- coding: utf-8 -*-
"""
Rango — Compras mensuales (Paradigma) → hoja 'Ordenado'.

Toma el 'Listado' de Libro IVA Compras que exporta el sistema de Rango y lo
reordena al formato 'Ordenado' que usa el estudio, agregando:
  - rubro contable (lookup por CUIT en el maestro de proveedores),
  - NG+NO GR = NG 21% + NG 10,5% + no gravados   (lo "amarillo"),
  - IVA TOTAL = IVA 21% + iva diferencial          (lo "amarillo"),
  - proceso = mes de proceso (se detecta del encabezado del Listado).
El resto de las columnas salen directo del Listado.
"""

import io
import re
from datetime import date, datetime

# Columnas de salida (hoja 'Ordenado'), en orden.
COLUMNAS = [
    "proceso", "Fecha", "Número", "Proveedor", "rubro contable", "CUIT", "IVA",
    "ORIGEN", "NG 21%", "NG 10,5%", "no Gravados", "NG+NO GR", "21 %", "iva dif",
    "IVA TOTAL", "Ret y Ret. IVA", "Ret y Ret. Ing. Brutos", "Ret y Ret. Ganancias",
    "Facturado",
]

_MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_cuit(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return "".join(ch for ch in s if ch.isdigit())


def _abrir(data):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


# --------------------------------------------------------------------------
# Maestro de proveedores de Rango (CUIT -> rubro contable + código cuenta)
# --------------------------------------------------------------------------
# Palabras clave -> rubro (sugerencia base, además de lo que se aprende del maestro).
_PALABRAS_RUBRO = [
    (("EST.DE SERVICIO", "EST. DE SERVICIO", "ESTACION DE SERVICIO", "ESTACION SERVICIO",
      "COMBUSTIBLE", "YPF", "SHELL", "AXION", "PUMA ENERGY", "GASOIL", "NAFTA"), "COMBUSTIBLES"),
    (("NEUMATIC", "GOMERIA", "CUBIERTA", " TIRE", "VULCA"), "REPUESTOS Y REPARACIONES"),
    (("REPUESTO", "AUTOPART", "MECANIC", "TALLER", "LUBRICANT"), "REPUESTOS Y REPARACIONES"),
    (("HIERRO", "ACERO", "CHAPA", "PERFIL", "ABERTURA", "GRIFERIA", "SANITARIO",
      "PINTURA", "CEMENTO", "CORRALON", "FERRETERIA", "MATERIALES", "HIMEBA",
      "ROTOPLAS", "WEBER", "CANNON"), "MERCADERIAS"),
    (("SEGURO", "ASEGURAD"), "SEGUROS"),
    (("TRANSPORTE", "FLETE", "LOGISTICA", "CARGAS", "EXPRESO"), "FLETES"),
    (("INMOBILIARIA", "ALQUILER"), "ALQUILERES"),
    (("COOP.ELECT", "COOP ELECT", "CAMUZZI", "AGUAS", "EDES", "EDEN", "ENERGIA",
      "ELECTRIC"), "LUZ-AGUA-GAS"),
    (("TELEFON", "MOVISTAR", "CLARO", "PERSONAL", "TELECOM", "FIBERTEL"), "TELEFONO"),
    (("BANCO", "TARJETA", "NARANJA", "VISA", "MASTERCARD"), "GASTOS BANCARIOS"),
    (("ESTUDIO", "CONTAD", "ABOGAD", "ESCRIBAN", "CONSULTOR"), "HONORARIOS Y ARANCELES"),
    (("LIMPIEZA",), "LIMPIEZA"),
    (("PUBLICIDAD", "RADIO", "DIARIO", "CARTEL"), "PUBLICIDAD"),
    (("COMPUTACION", "SISTEMAS", "SOFTWARE", "INFORMATICA", "PARADIGMA",
      "NIC ARGENTINA"), "GASTOS COMPUTACION"),
    (("PAPELERIA", "LIBRERIA", "IMPRENTA"), "PAPELERIA E INSUMOS"),
]

_STOP = {"SRL", "S.R.L.", "SA", "S.A.", "SACIF", "SAIC", "SACI", "DE", "DEL", "LA",
         "EL", "LOS", "LAS", "Y", "E", "S.A", "S.R.L", "CIA", "HNOS", "SAS"}


def _tokens(nombre):
    import re
    bruto = re.split(r"[^0-9A-Za-zÁÉÍÓÚÑ]+", str(nombre).upper())
    return [t for t in bruto if len(t) >= 4 and t not in _STOP]


def aprender_rubros(pares):
    """pares: lista de (nombre, rubro). Devuelve token -> Counter(rubro)."""
    from collections import defaultdict, Counter
    mapa = defaultdict(Counter)
    for nombre, rubro in pares:
        if not rubro:
            continue
        for t in _tokens(nombre):
            mapa[t][rubro] += 1
    return mapa


def sugerir_rubro(nombre, aprendido, rubros_validos):
    """Sugiere un rubro para un proveedor nuevo. Primero por palabras clave fuertes
    (alta precisión); si no, por lo aprendido del maestro pero SOLO con tokens
    'limpios' (un rubro claramente dominante). Devuelve '' si no hay señal clara."""
    txt = " " + str(nombre).upper() + " "
    for claves, rubro in _PALABRAS_RUBRO:
        if any(k in txt for k in claves):
            if not rubros_validos or rubro in rubros_validos:
                return rubro

    from collections import Counter
    puntajes = Counter()
    for t in _tokens(nombre):
        cnt = aprendido.get(t)
        if not cnt:
            continue
        total = sum(cnt.values())
        rubro, n = cnt.most_common(1)[0]
        # solo si ese token apunta claramente a un rubro (evita apellidos genéricos)
        if n >= 2 and n / total >= 0.6:
            puntajes[rubro] += n
    if puntajes:
        mejor = puntajes.most_common(1)[0][0]
        if not rubros_validos or mejor in rubros_validos:
            return mejor
    return ""


def leer_maestro_pares(data):
    """Devuelve lista de (Razón Social, Rubro) del maestro, para aprender."""
    wb = _abrir(data)
    hoja = next((h for h in wb.sheetnames
                 if "proveedor" in h.lower() or "rubros-prov" in h.lower()), wb.sheetnames[0])
    ws = wb[hoja]
    hdr = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def idx(*claves):
        for i, h in enumerate(hdr):
            if any(k in h for k in claves):
                return i + 1
        return None

    c_nom = idx("raz", "proveedor", "nombre") or 2
    c_rubro = idx("rubro") or 3
    pares = []
    for r in range(2, ws.max_row + 1):
        nom = ws.cell(r, c_nom).value
        rub = ws.cell(r, c_rubro).value
        if nom and rub:
            pares.append((str(nom), str(rub).strip()))
    return pares


def leer_maestro(data):
    """Devuelve (proveedores: CUIT->(rubro, codigo), rubros: set de rubros)."""
    wb = _abrir(data)
    hoja = next((h for h in wb.sheetnames
                 if "proveedor" in h.lower() or "rubros-prov" in h.lower()), None)
    if hoja is None:
        hoja = wb.sheetnames[0]
    ws = wb[hoja]
    hdr = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def idx(*claves):
        for i, h in enumerate(hdr):
            if any(k in h for k in claves):
                return i + 1
        return None

    c_cuit = idx("cuit") or 1
    c_rubro = idx("rubro") or 3
    c_cod = idx("codigo", "código", "cuenta")
    proveedores, rubros = {}, set()
    rubro_codigo = {}
    for r in range(2, ws.max_row + 1):
        cuit = _norm_cuit(ws.cell(r, c_cuit).value)
        rubro = ws.cell(r, c_rubro).value
        cod = ws.cell(r, c_cod).value if c_cod else ""
        if cuit and rubro not in (None, ""):
            rb = str(rubro).strip()
            cd = str(cod or "").strip()
            proveedores[cuit] = (rb, cd)
            rubros.add(rb)
            if cd and rb not in rubro_codigo:
                rubro_codigo[rb] = cd
    return proveedores, rubros, rubro_codigo


# --------------------------------------------------------------------------
# Listado crudo de Paradigma
# --------------------------------------------------------------------------
def detectar_proceso(ws):
    """Lee el rango de fechas del encabezado y devuelve date(año, mes, 1)."""
    for r in range(1, 12):
        for c in range(1, 4):
            txt = str(ws.cell(r, c).value or "").lower()
            m = re.search(r"de\s+([a-záéíóú]+)\s+de\s+(\d{4})", txt)
            if m and m.group(1) in _MESES:
                return date(int(m.group(2)), _MESES[m.group(1)], 1)
    return None


def _fila_encabezado(ws):
    """Encuentra la fila de títulos del Listado (la que tiene CUIT y Fecha)."""
    for r in range(1, 20):
        fila = " ".join(str(ws.cell(r, c).value or "").lower() for c in range(1, ws.max_column + 1))
        if "cuit" in fila and "fecha" in fila:
            return r
    return 8


def procesar(listado_bytes, maestro_bytes, proceso=None):
    """Devuelve (filas, desconocidos, stats, proceso_detectado, rubros)."""
    wb = _abrir(listado_bytes)
    ws = wb[wb.sheetnames[0]]
    proceso_det = detectar_proceso(ws)
    if proceso is None:
        proceso = proceso_det

    hr = _fila_encabezado(ws)
    hdr = [str(ws.cell(hr, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def col(*claves):
        for i, h in enumerate(hdr):
            if any(k in h for k in claves):
                return i + 1
        return None

    c_fecha = col("fecha")
    c_num = col("número", "numero", "compr")
    c_nom = col("nombre")
    c_cuit = col("cuit")
    c_iva = col("iva") if hdr.count("iva") else None
    # 'IVA' (condición) es la columna corta; buscar exacta
    c_cond = next((i + 1 for i, h in enumerate(hdr) if h == "iva"), None)
    c_orig = col("tipoimputacion", "imputaci")
    c_n21 = col("neto 21")
    c_ndif = col("neto diferencial", "diferencial")
    c_ng = col("no gravado", "conceptos no")
    c_i21 = next((i + 1 for i, h in enumerate(hdr) if h.startswith("iva 21")), None)
    c_idif = next((i + 1 for i, h in enumerate(hdr) if "iva dif" in h or h == "iva diferencial"), None)
    c_riva = col("rets. iva", "rets iva", "rets. iva")
    c_riibb = col("iibb", "ing. brutos", "ingresos brutos")
    c_rgan = col("ganancias")
    c_total = col("imp. total", "imp total", "total")

    proveedores, rubros, rubro_codigo = leer_maestro(maestro_bytes)

    filas = []
    desconocidos = {}
    asignados = 0
    _MESNOM = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
               "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    proceso_txt = f"{_MESNOM[proceso.month]} {proceso.year}" if proceso else ""
    for r in range(hr + 1, ws.max_row + 1):
        # cortar en 'Totales:' o fila sin fecha
        marca = " ".join(str(ws.cell(r, c).value or "").lower() for c in range(1, ws.max_column + 1))
        if "totales" in marca:
            break
        fecha = ws.cell(r, c_fecha).value if c_fecha else None
        cuit = _norm_cuit(ws.cell(r, c_cuit).value) if c_cuit else ""
        if (fecha in (None, "")) and not cuit:
            continue

        rubro, _cod = proveedores.get(cuit, ("", ""))
        if rubro == "":
            nom = ws.cell(r, c_nom).value if c_nom else ""
            if cuit:
                desconocidos.setdefault(cuit, nom)
        else:
            asignados += 1

        ng21 = _num(ws.cell(r, c_n21).value) if c_n21 else 0.0
        ng105 = _num(ws.cell(r, c_ndif).value) if c_ndif else 0.0
        nogr = _num(ws.cell(r, c_ng).value) if c_ng else 0.0
        iva21 = _num(ws.cell(r, c_i21).value) if c_i21 else 0.0
        ivadif = _num(ws.cell(r, c_idif).value) if c_idif else 0.0

        filas.append({
            "proceso": proceso_txt,
            "Fecha": ws.cell(r, c_fecha).value if c_fecha else "",
            "Número": ws.cell(r, c_num).value if c_num else "",
            "Proveedor": ws.cell(r, c_nom).value if c_nom else "",
            "rubro contable": rubro,
            "CUIT": cuit,
            "IVA": ws.cell(r, c_cond).value if c_cond else "",
            "ORIGEN": ws.cell(r, c_orig).value if c_orig else "",
            "NG 21%": ng21,
            "NG 10,5%": ng105,
            "no Gravados": nogr,
            "NG+NO GR": round(ng21 + ng105 + nogr, 2),
            "21 %": iva21,
            "iva dif": ivadif,
            "IVA TOTAL": round(iva21 + ivadif, 2),
            "Ret y Ret. IVA": _num(ws.cell(r, c_riva).value) if c_riva else 0.0,
            "Ret y Ret. Ing. Brutos": _num(ws.cell(r, c_riibb).value) if c_riibb else 0.0,
            "Ret y Ret. Ganancias": _num(ws.cell(r, c_rgan).value) if c_rgan else 0.0,
            "Facturado": _num(ws.cell(r, c_total).value) if c_total else 0.0,
        })

    stats = {
        "comprobantes": len(filas),
        "asignados": asignados,
        "sin_rubro": len(filas) - asignados,
        "proveedores_nuevos": len(desconocidos),
    }
    return filas, desconocidos, stats, proceso_det, sorted(rubros), rubro_codigo


# --------------------------------------------------------------------------
# Salidas
# --------------------------------------------------------------------------
def _es_bancario(rubro):
    return "BANCAR" in str(rubro).upper()


def _proc_key(texto):
    """'diciembre 2025' -> (2025, 12) para ordenar cronológicamente."""
    p = str(texto).split()
    try:
        mes = _MESNOM_L.index(p[0].lower())
        return (int(p[1]), mes)
    except (ValueError, IndexError):
        return (9999, 99)


_MESNOM_L = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
             "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def leer_ordenado(data):
    """Lee la hoja 'Ordenado' de un archivo acumulado y devuelve la lista de filas
    (dicts con COLUMNAS), para poder sumarle un mes nuevo."""
    wb = _abrir(data)
    hoja = next((h for h in wb.sheetnames if h.strip().lower() == "ordenado"), None)
    if hoja is None:
        return []
    ws = wb[hoja]
    # ubicar la columna donde arranca el encabezado (puede haber col A vacía)
    cab = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            cab[str(v).strip()] = c
    if "proceso" not in cab:
        return []
    filas = []
    for r in range(2, ws.max_row + 1):
        if not any(ws.cell(r, c).value not in (None, "") for c in cab.values()):
            continue
        fila = {}
        for col in COLUMNAS:
            c = cab.get(col)
            fila[col] = ws.cell(r, c).value if c else ("" if col in ("proceso", "Fecha", "Número", "Proveedor", "rubro contable", "CUIT", "IVA", "ORIGEN") else 0.0)
        if not str(fila.get("proceso") or "").strip():
            continue
        filas.append(fila)
    return filas


def construir_libro(filas, maestro_bytes=None, rubro_codigo=None):
    """Arma el libro de un mes: Ordenado (Tabla de Excel) + iva+retenciones +
    TD-RUBROS-ASIENTOS (el asiento completo: rubros + gastos bancarios + tarjetas
    a cobrar + IVA + retenciones + 'a caja') + TARJETAS + RESUMEN (si el maestro
    la trae, con COMPRAS = mercaderías del mes)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    rubro_codigo = rubro_codigo or {}
    bold = Font(bold=True)
    amar = PatternFill("solid", fgColor="FFFF00")
    NUM = "#,##0.00"
    mes_lbl = str(filas[0]["proceso"]).strip() if filas else "mes"

    def suma(clave, filt=None):
        return round(sum(f[clave] for f in filas if (filt is None or filt(f))), 2)

    wb = openpyxl.Workbook()

    # ---------- Hoja Ordenado (como Tabla de Excel, para las TD reales) ----------
    ws = wb.active
    ws.title = "Ordenado"
    amarillas = {"rubro contable", "NG+NO GR", "IVA TOTAL"}
    for j, t in enumerate(COLUMNAS, start=2):
        c = ws.cell(1, j, t); c.font = bold
        if t in amarillas:
            c.fill = amar
    for i, fila in enumerate(filas, start=2):
        for j, t in enumerate(COLUMNAS, start=2):
            c = ws.cell(i, j, fila.get(t))
            if isinstance(fila.get(t), (int, float)) and t != "CUIT":
                c.number_format = NUM
            if t in amarillas:
                c.fill = amar
    ws.freeze_panes = "C2"
    for col, w in {"C": 12, "D": 18, "E": 32, "F": 22, "G": 14}.items():
        ws.column_dimensions[col].width = w
    ult = len(filas) + 1
    if ult >= 2:
        ref = f"B1:{openpyxl.utils.get_column_letter(1 + len(COLUMNAS))}{ult}"
        tab = Table(displayName="Ordenado", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
        ws.add_table(tab)

    # ---------- iva+retenciones ----------
    wi = wb.create_sheet("iva+retenciones")
    wi.cell(1, 2, "Datos").font = bold
    wi.cell(1, 3, mes_lbl).font = bold
    for i, m in enumerate(["NG+NO GR", "IVA TOTAL", "Ret y Ret. IVA",
                           "Ret y Ret. Ing. Brutos", "Ret y Ret. Ganancias"], start=2):
        wi.cell(i, 2, "Suma de " + m)
        wi.cell(i, 3, suma(m)).number_format = NUM
    wi.column_dimensions["B"].width = 30

    # ---------- TARJETAS (Gastos Bancarios del mes) ----------
    wj = wb.create_sheet("TARJETAS")
    wj.cell(2, 2, "rubro contable").font = bold
    wj.cell(2, 3, rubro_codigo.get("GASTOS BANCARIOS", "53900000-GASTOS BANCARIOS"))
    wj.cell(4, 2, "Valores").font = bold
    wj.cell(4, 3, mes_lbl).font = bold
    banc = lambda f: _es_bancario(f["rubro contable"])
    t_nograv = suma("no Gravados", banc)
    t_ng21 = suma("NG 21%", banc)
    t_ng105 = suma("NG 10,5%", banc)
    for i, (m, v) in enumerate([("no Gravados", t_nograv), ("NG 21%", t_ng21),
                                ("NG 10,5%", t_ng105)], start=5):
        wj.cell(i, 2, "Suma de " + m)
        wj.cell(i, 3, v).number_format = NUM
    wj.cell(8, 3, round(t_nograv + t_ng21 + t_ng105, 2)).number_format = NUM
    wj.cell(8, 3).font = bold
    wj.column_dimensions["B"].width = 22

    # ---------- TD-RUBROS-ASIENTOS (el asiento completo del mes) ----------
    wt = wb.create_sheet("TD-RUBROS-ASIENTOS")
    wt.cell(1, 2, "(excluir GASTOS BANCARIOS)").font = bold
    wt.cell(2, 2, "rubro contable").font = bold
    wt.cell(2, 3, mes_lbl).font = bold
    porrub = {}
    for f in filas:
        if _es_bancario(f["rubro contable"]):
            continue
        rb = f["rubro contable"] or "(sin rubro)"
        porrub[rb] = porrub.get(rb, 0) + f["NG+NO GR"]
    r = 3
    total = 0.0
    for rb in sorted(porrub):
        wt.cell(r, 2, rubro_codigo.get(rb, rb))
        c = wt.cell(r, 3, round(porrub[rb], 2)); c.number_format = NUM
        total += porrub[rb]; r += 1
    # Gastos bancarios (parte gravada de tarjetas) y tarjetas a cobrar (no gravado)
    cod_banc = rubro_codigo.get("GASTOS BANCARIOS", "53900000-GASTOS BANCARIOS")
    cb = wt.cell(r, 2, cod_banc); cb.fill = amar
    c = wt.cell(r, 3, round(t_ng21 + t_ng105, 2)); c.number_format = NUM; c.fill = amar
    wt.cell(r, 4, "ng 21% + ng 10,5% de la hoja TARJETAS")
    total += t_ng21 + t_ng105; r += 1
    cb = wt.cell(r, 2, "TARJETAS A COBRAR (NETO ACREDITADO)"); cb.fill = amar
    c = wt.cell(r, 3, round(t_nograv, 2)); c.number_format = NUM; c.fill = amar
    wt.cell(r, 4, "no gravado de la hoja TARJETAS")
    total += t_nograv; r += 1
    # IVA y retenciones
    for m in ["IVA TOTAL", "Ret y Ret. IVA", "Ret y Ret. Ing. Brutos", "Ret y Ret. Ganancias"]:
        wt.cell(r, 2, "Suma de " + m)
        v = suma(m)
        wt.cell(r, 3, v).number_format = NUM
        total += v; r += 1
    # Total y 'a caja'
    ctot = wt.cell(r, 3, round(total, 2)); ctot.number_format = NUM; ctot.font = bold
    r += 1
    wt.cell(r, 2, "a caja").font = bold
    cc = wt.cell(r, 3, round(total, 2)); cc.number_format = NUM; cc.font = bold
    wt.column_dimensions["B"].width = 38
    wt.column_dimensions["C"].width = 16
    wt.column_dimensions["D"].width = 34

    # ---------- RESUMEN (si el maestro la trae) ----------
    if maestro_bytes:
        try:
            src = openpyxl.load_workbook(io.BytesIO(maestro_bytes), data_only=True)
            if "RESUMEN" in src.sheetnames:
                rs = src["RESUMEN"]
                wr = wb.create_sheet("RESUMEN")
                compras = suma("NG+NO GR", lambda f: "MERCADER" in str(f["rubro contable"]).upper())
                for rr in range(1, rs.max_row + 1):
                    for cc2 in range(1, rs.max_column + 1):
                        v = rs.cell(rr, cc2).value
                        if v is not None:
                            celda = wr.cell(rr, cc2, v)
                            if isinstance(v, (int, float)):
                                celda.number_format = NUM
                        if str(rs.cell(rr, 1).value).strip().upper() == "COMPRAS" and cc2 == 2:
                            x = wr.cell(rr, cc2, compras); x.number_format = NUM; x.fill = amar
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_excel(filas):
    return construir_libro(filas)


RUBROS_RANGO = [
    "ALQUILERES", "COMBUSTIBLES", "FLETES", "GASTOS BANCARIOS", "GASTOS COMPUTACION",
    "GASTOS VARIOS", "HONORARIOS Y ARANCELES", "INSTALACIONES Y MEJORAS", "LIMPIEZA",
    "LUZ-AGUA-GAS", "MERCADERIAS", "PAPELERIA E INSUMOS", "PUBLICIDAD",
    "REPUESTOS Y REPARACIONES", "RODADOS", "RODADOS V. ORIGEN", "SEGUROS",
    "TELEFONO", "VEHICULOS",
]


def construir_plantilla_maestro(rubros=None):
    """Excel maestro vacío para Rango: Proveedores (encabezados) + Rubros + Instrucciones."""
    import openpyxl
    from openpyxl.styles import Font
    rubros = rubros or RUBROS_RANGO
    bold = Font(bold=True)
    wb = openpyxl.Workbook()
    wi = wb.active
    wi.title = "Instrucciones"
    for i, t in enumerate([
        "MAESTRO DE PROVEEDORES - RANGO (Compras Paradigma)",
        "",
        "Hoja 'Proveedores': cargá cada proveedor con su CUIT, nombre, rubro contable",
        "y (opcional) el código de cuenta. La herramienta busca el rubro por CUIT.",
        "Hoja 'Rubros': lista de rubros contables disponibles (editable).",
        "",
        "Tip: en la herramienta, cuando aparezca un proveedor nuevo lo cargás ahí y",
        "te devuelve este maestro actualizado.",
    ], start=1):
        wi.cell(i, 1, t)
    wp = wb.create_sheet("Proveedores")
    for j, t in enumerate(["CUIT", "Razón Social", "Rubro", "Código cuenta contable"], start=1):
        wp.cell(1, j, t).font = bold
    wp.column_dimensions["A"].width = 16
    wp.column_dimensions["B"].width = 38
    wp.column_dimensions["D"].width = 30
    wr = wb.create_sheet("Rubros")
    wr.cell(1, 1, "Rubro contable").font = bold
    for i, rb in enumerate(rubros, start=2):
        wr.cell(i, 1, rb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_maestro_actualizado(maestro_bytes, nuevos):
    """nuevos = lista de (cuit, proveedor, rubro, codigo)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(maestro_bytes))
    hoja = next((h for h in wb.sheetnames
                 if "proveedor" in h.lower() or "rubros-prov" in h.lower()), wb.sheetnames[0])
    ws = wb[hoja]
    fila = ws.max_row + 1
    for cuit, prov, rubro, cod in nuevos:
        ws.cell(fila, 1, str(cuit))
        ws.cell(fila, 2, prov)
        ws.cell(fila, 3, rubro)
        ws.cell(fila, 4, cod)
        fila += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
