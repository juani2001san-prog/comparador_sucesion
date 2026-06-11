# -*- coding: utf-8 -*-
"""
Ventas por actividad (Tango) para la DJ de IVA y Convenio Multilateral.

Toma el export "IVA POR ACTIVIDAD" de Tango (un reporte en árbol:
Concepto -> Actividad -> Clasificación -> Categoría de IVA -> alícuota) y arma:

  - una tabla PLANA (un renglón por hoja del árbol),
  - una tabla por ACTIVIDAD x CATEGORÍA de IVA separando FACTURA y NOTA DE
    CRÉDITO, con el Total neto = ventas - NC,
  - un RESUMEN por actividad (RI / Monotributo / CF) con neto e IVA.

Reglas (deducidas del armado manual del usuario):
  - Concepto 01 (IVA POR ACTIVIDAD)            -> FACTURA
  - Concepto 02 (RESTITUCIÓN DE DÉBITO FISCAL) -> NOTA DE CRÉDITO
  - El neto y el IVA se calculan SIEMPRE desde el Importe Facturado:
        NETO = Facturado / (1 + alícuota)
        IVA  = Facturado - NETO
  - En Nota de Crédito, neto e IVA van en NEGATIVO (restan).
"""

import io
import re

import pandas as pd


# --------------------------------------------------------------------------
# Lectura robusta (.xls viejo via xlrd / .xlsx via openpyxl)
# --------------------------------------------------------------------------
def _leer_matriz(data):
    """Devuelve la primera hoja como lista de filas (listas de celdas)."""
    if data[:2] == b"PK":  # xlsx (zip)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        return [list(fila) for fila in ws.iter_rows(values_only=True)]
    # .xls viejo (OLE2)
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    sh = wb.sheet_by_index(0)
    return [[sh.cell(r, c).value for c in range(sh.ncols)] for r in range(sh.nrows)]


def _num(valor):
    """Parsea un importe que puede venir como número o texto '136,823,564.78'."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace(" ", "")
    if s.endswith("%"):
        s = s[:-1]
    s = s.replace(",", "")  # separador de miles (formato del export)
    try:
        return float(s)
    except ValueError:
        return None


def _alicuota(texto):
    """De '21.00%' -> 0.21 ; '0.00%' -> 0.0 ; '10.50%' -> 0.105."""
    s = str(texto).strip()
    if not s.endswith("%"):
        return None
    v = _num(s)
    return None if v is None else round(v / 100.0, 5)


def _categoria_corta(texto):
    t = str(texto).upper()
    if "RESPONSABLE" in t:
        return "RI"
    if "MONOTRIB" in t:
        return "Monot."
    if "C.F" in t or "EX" in t or "NO ALC" in t or "CONSUMIDOR" in t:
        return "CF, Ex., No alc."
    return str(texto).strip()


# --------------------------------------------------------------------------
# Parseo del árbol de Tango -> tabla plana
# --------------------------------------------------------------------------
def _tabla_plana(data):
    rows = _leer_matriz(data)

    # Encontrar la fila de encabezados (la que tiene 'facturado' y 'neto').
    hdr = None
    for i, r in enumerate(rows):
        txt = " ".join(str(x) for x in r if x).lower()
        if "facturado" in txt and "neto" in txt:
            hdr = i
            break
    if hdr is None:
        raise ValueError("No encontré el encabezado del reporte de Tango "
                         "(faltan las columnas 'Importe neto' / 'Importe facturado').")

    header = rows[hdr]

    def col(clave):
        for j, x in enumerate(header):
            if x and clave in str(x).lower():
                return j
        return None

    c_fact = col("facturado")
    if c_fact is None:
        raise ValueError("No encontré la columna 'Importe facturado'.")

    concepto = actividad = categoria = None
    registros = []

    for r in rows[hdr + 1:]:
        celdas = ["" if x is None else x for x in r]

        # Actualizar los niveles del árbol según las etiquetas.
        for x in celdas:
            s = str(x)
            if "Concepto :" in s:
                concepto = s.split("Concepto :", 1)[1].strip()
            elif "Actividad :" in s:
                actividad = s.split("Actividad :", 1)[1].strip()
            elif ("Categor" in s) and ("I.V.A" in s):
                categoria = s.rsplit(":", 1)[-1].strip()

        # ¿Es una hoja con importe? (tiene Facturado numérico + una alícuota)
        fact = _num(celdas[c_fact]) if c_fact < len(celdas) else None
        alic = None
        for x in celdas:
            a = _alicuota(x)
            if a is not None:
                alic = a
                break

        if fact is None or alic is None or not actividad or not categoria:
            continue
        if "TOTAL" in " ".join(str(x) for x in celdas).upper():
            continue  # saltear subtotales si los hubiera

        es_nc = concepto.strip().startswith("02") or "RESTITU" in concepto.upper()
        comp = "NOTA CREDITO" if es_nc else "FACTURA"
        neto = fact / (1 + alic)
        iva = fact - neto
        signo = -1 if es_nc else 1

        registros.append({
            "Comprobante": comp,
            "Actividad": actividad,
            "Categoria": categoria,
            "Cat": _categoria_corta(categoria),
            "Alicuota": alic,
            "Facturado": round(signo * fact, 2),
            "Neto": round(signo * neto, 2),
            "IVA": round(signo * iva, 2),
        })

    if not registros:
        raise ValueError("No se pudo extraer ningún movimiento. ¿Es el reporte "
                         "'IVA por actividad' de Tango?")
    return pd.DataFrame(registros)


# --------------------------------------------------------------------------
# Agregaciones
# --------------------------------------------------------------------------
def _por_actividad_categoria(df):
    """Actividad x Categoría: NC, Factura y Total (= Factura - NC) de neto e IVA."""
    g = df.groupby(["Actividad", "Categoria", "Comprobante"])[["Neto", "IVA"]].sum()
    filas = []
    for (act, cat), sub in g.groupby(level=[0, 1]):
        def val(comp, col):
            try:
                return round(float(sub.loc[(act, cat, comp), col]), 2)
            except KeyError:
                return 0.0
        nc_neto, nc_iva = val("NOTA CREDITO", "Neto"), val("NOTA CREDITO", "IVA")
        fa_neto, fa_iva = val("FACTURA", "Neto"), val("FACTURA", "IVA")
        filas.append({
            "Actividad": act,
            "Categoría de IVA": cat,
            "_o": _orden_cat(_categoria_corta(cat)),
            "NC Neto": nc_neto, "NC IVA": nc_iva,
            "Factura Neto": fa_neto, "Factura IVA": fa_iva,
            "Total Neto": round(fa_neto + nc_neto, 2),
            "Total IVA": round(fa_iva + nc_iva, 2),
        })
    out = (pd.DataFrame(filas).sort_values(["Actividad", "_o"])
           .drop(columns="_o").reset_index(drop=True))
    return out


_ORDEN_CAT = {"RI": 0, "Monot.": 1, "CF, Ex., No alc.": 2}


def _orden_cat(cat_corta):
    return _ORDEN_CAT.get(cat_corta, 9)


def _resumen(df):
    """Resumen por actividad y categoría (neto e IVA netos = ventas - NC)."""
    g = df.groupby(["Actividad", "Cat"])[["Neto", "IVA"]].sum().round(2).reset_index()
    g = g.rename(columns={"Cat": "Categoría", "Neto": "Neto Grav.+Ex.", "IVA": "IVA"})
    g["_o"] = g["Categoría"].map(_orden_cat)
    g = g.sort_values(["Actividad", "_o"]).drop(columns="_o").reset_index(drop=True)
    return g


def procesar(data):
    """Devuelve (detalle, por_actividad, resumen, totales)."""
    detalle = _tabla_plana(data)
    por_act = _por_actividad_categoria(detalle)
    resumen = _resumen(detalle)
    totales = {
        "Neto": round(float(detalle["Neto"].sum()), 2),
        "IVA": round(float(detalle["IVA"].sum()), 2),
        "Facturado": round(float(detalle["Facturado"].sum()), 2),
    }
    return detalle, por_act, resumen, totales


# --------------------------------------------------------------------------
# Excel de salida (detallado, estilo armado manual)
# --------------------------------------------------------------------------
def construir_excel(detalle, por_act, resumen, totales):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Por actividad"

    bold = Font(bold=True)
    titulo = Font(bold=True, size=13)
    centro = Alignment(horizontal="center")
    gris = PatternFill("solid", fgColor="D9E1F2")
    amar = PatternFill("solid", fgColor="FFF2CC")
    fino = Side(style="thin", color="BFBFBF")
    borde = Border(left=fino, right=fino, top=fino, bottom=fino)
    NUM = "#,##0.00"

    def celda(r, c, v, *, font=None, fill=None, fmt=None, al=None, bd=True):
        cell = ws.cell(r, c, v)
        if font: cell.font = font
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt
        if al: cell.alignment = al
        if bd: cell.border = borde
        return cell

    ws.cell(1, 1, "VENTAS POR ACTIVIDAD  (ventas − notas de crédito)").font = titulo

    # --- Encabezado de la tabla (fila 3 grupos, fila 4 columnas) ---
    # B=Actividad C=Cliente | D,E=NC | F,G=Factura | H,I=Total
    celda(3, 4, "NOTA DE CRÉDITO", font=bold, fill=gris, al=centro)
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
    celda(3, 5, None, fill=gris)
    celda(3, 6, "FACTURA", font=bold, fill=gris, al=centro)
    ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)
    celda(3, 7, None, fill=gris)
    celda(3, 8, "TOTAL (Vta − NC)", font=bold, fill=amar, al=centro)
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=9)
    celda(3, 9, None, fill=amar)

    enc = ["ACTIVIDAD", "CLIENTE (Cat. IVA)", "Neto", "IVA", "Neto", "IVA", "Neto", "IVA"]
    for j, t in enumerate(enc, start=2):
        celda(4, j, t, font=bold, fill=(amar if j >= 8 else gris), al=centro)

    # --- Filas de datos ---
    r = 5
    act_prev = None
    tot = {"d": 0, "e": 0, "f": 0, "g": 0, "h": 0, "i": 0}
    for _, row in por_act.iterrows():
        act = row["Actividad"]
        celda(r, 2, "" if act == act_prev else act)
        act_prev = act
        celda(r, 3, row["Categoría de IVA"])
        vals = [row["NC Neto"], row["NC IVA"], row["Factura Neto"], row["Factura IVA"],
                row["Total Neto"], row["Total IVA"]]
        for k, (col, v) in enumerate(zip("defghi", vals)):
            celda(r, 4 + k, v, fmt=NUM, fill=(amar if 4 + k >= 8 else None))
            tot[col] += v
        r += 1

    # --- Total general ---
    celda(r, 2, "Total general", font=bold)
    celda(r, 3, "", font=bold)
    for k, col in enumerate("defghi"):
        celda(r, 4 + k, round(tot[col], 2), font=bold, fmt=NUM,
              fill=(amar if 4 + k >= 8 else gris))
    r += 3

    # --- Resumen por actividad ---
    celda(r, 2, "RESUMEN POR ACTIVIDAD", font=titulo, bd=False)
    r += 1
    celda(r, 2, "Actividad", font=bold, fill=gris)
    celda(r, 3, "Cliente (Cat. IVA)", font=bold, fill=gris)
    celda(r, 4, "Neto Grav.+Ex.", font=bold, fill=gris, al=centro)
    celda(r, 5, "IVA", font=bold, fill=gris, al=centro)
    r += 1
    suma_neto = suma_iva = 0.0
    for act, sub in resumen.groupby("Actividad", sort=False):
        celda(r, 2, act, font=bold)
        celda(r, 3, "", bd=True); celda(r, 4, None, bd=True); celda(r, 5, None, bd=True)
        r += 1
        for _, row in sub.iterrows():
            celda(r, 2, "")
            celda(r, 3, row["Categoría"])
            celda(r, 4, row["Neto Grav.+Ex."], fmt=NUM)
            celda(r, 5, row["IVA"], fmt=NUM)
            suma_neto += float(row["Neto Grav.+Ex."]); suma_iva += float(row["IVA"])
            r += 1
    # Totales y control
    r += 1
    celda(r, 3, "Neto Grav.+Ex.", font=bold, fill=amar, al=centro)
    celda(r, 4, "IVA", font=bold, fill=amar, al=centro)
    celda(r, 5, "", fill=amar)
    r += 1
    celda(r, 2, "TOTAL", font=bold)
    celda(r, 3, round(suma_neto, 2), font=bold, fmt=NUM, fill=amar)
    celda(r, 4, round(suma_iva, 2), font=bold, fmt=NUM, fill=amar)
    celda(r, 5, "", fill=amar)
    r += 1
    dif_neto = round(suma_neto - tot["h"], 2)
    dif_iva = round(suma_iva - tot["i"], 2)
    celda(r, 2, "DIF (control = 0)", font=bold)
    celda(r, 3, dif_neto, fmt=NUM)
    celda(r, 4, dif_iva, fmt=NUM)

    # Anchos de columna
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 26
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 16

    # --- Hoja Detalle ---
    wd = wb.create_sheet("Detalle")
    cols = ["Comprobante", "Actividad", "Categoria", "Alicuota", "Facturado", "Neto", "IVA"]
    for j, t in enumerate(cols, start=1):
        wd.cell(1, j, t).font = bold
    for i, (_, row) in enumerate(detalle.iterrows(), start=2):
        for j, t in enumerate(cols, start=1):
            c = wd.cell(i, j, row[t])
            if t in ("Facturado", "Neto", "IVA"):
                c.number_format = NUM
            if t == "Alicuota":
                c.number_format = "0.0%"
    wd.column_dimensions["B"].width = 46
    wd.column_dimensions["C"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
